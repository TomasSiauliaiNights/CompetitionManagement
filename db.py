"""
Database module for Tournament Manager
- PostgreSQL as primary real-time store
- Excel import for initial robot loading
- Optional Google Sheets background sync
"""

import threading, time, traceback

try:
    import psycopg2
    import psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False

try:
    import openpyxl
    HAS_XL = True
except ImportError:
    HAS_XL = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    HAS_GS = True
except ImportError:
    HAS_GS = False


def parse_time(s):
    """Parse '0:07.699' or '7.699' → milliseconds."""
    if s is None or str(s).strip().upper() in ("DNF", ""):
        return None
    s = str(s).strip()
    try:
        if ":" in s:
            p = s.split(":")
            sp = p[1].split(".")
            return int(p[0]) * 60000 + int(sp[0]) * 1000 + (int(sp[1].ljust(3, "0")[:3]) if len(sp) > 1 else 0)
        sp = s.split(".")
        return int(sp[0]) * 1000 + (int(sp[1].ljust(3, "0")[:3]) if len(sp) > 1 else 0)
    except:
        return None


class DB:
    """PostgreSQL tournament database."""

    def __init__(self):
        self.conn = None
        self.dsn = None

    @property
    def ok(self):
        return self.conn is not None and not self.conn.closed

    def connect(self, host="localhost", port=5432, dbname="tournament", user="postgres", password=""):
        if not HAS_PG:
            raise ImportError("pip install psycopg2-binary")
        self.dsn = f"host={host} port={port} dbname={dbname} user={user} password={password}"
        self.conn = psycopg2.connect(self.dsn)
        self.conn.autocommit = True
        self._ensure_schema()
        print(f"[DB] Connected to {dbname}@{host}:{port}")

    def _ensure_schema(self):
        """Create tables if they don't exist."""
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS robots (
                id SERIAL PRIMARY KEY,
                number VARCHAR(20) NOT NULL,
                name VARCHAR(100) NOT NULL DEFAULT '',
                category VARCHAR(30) NOT NULL,
                inspection BOOLEAN DEFAULT FALSE,
                UNIQUE(number, category)
            );
            CREATE TABLE IF NOT EXISTS trials (
                id SERIAL PRIMARY KEY,
                robot_id INTEGER REFERENCES robots(id) ON DELETE CASCADE,
                trial_num INTEGER NOT NULL,
                value VARCHAR(50),
                value_ms INTEGER,
                value_points INTEGER,
                is_dnf BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS folkrace_groups (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                sort_order INTEGER DEFAULT 0,
                section VARCHAR(10) DEFAULT 'L'
            );
            CREATE TABLE IF NOT EXISTS folkrace_entries (
                id SERIAL PRIMARY KEY,
                group_id INTEGER REFERENCES folkrace_groups(id) ON DELETE CASCADE,
                robot_id INTEGER REFERENCES robots(id) ON DELETE CASCADE,
                r1 INTEGER DEFAULT 0,
                r2 INTEGER DEFAULT 0,
                r3 INTEGER DEFAULT 0,
                total INTEGER DEFAULT 0
            );
        """)
        cur.close()

    def _q(self, sql, params=None):
        cur = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params or ())
        try:
            return cur.fetchall()
        except:
            return []
        finally:
            cur.close()

    def _x(self, sql, params=None):
        cur = self.conn.cursor()
        cur.execute(sql, params or ())
        cur.close()

    # ── Robots ──

    def add_robot(self, number, name, category):
        self._x("""
            INSERT INTO robots (number, name, category)
            VALUES (%s, %s, %s)
            ON CONFLICT (number, category) DO UPDATE SET name = EXCLUDED.name
        """, (str(number).strip(), name.strip(), category))

    def get_robots(self, category):
        return self._q(
            "SELECT id, number, name FROM robots WHERE category = %s ORDER BY number",
            (category,))

    def find_robot(self, number, category):
        rows = self._q(
            "SELECT id, number, name FROM robots WHERE number = %s AND category = %s",
            (str(number).strip(), category))
        return rows[0] if rows else None

    # ── Trials (Line Following / Fire Sister) ──

    def get_trials(self, robot_id, max_trials=5):
        """Returns list of trial dicts ordered by trial_num."""
        return self._q(
            "SELECT trial_num, value, value_ms, value_points, is_dnf FROM trials "
            "WHERE robot_id = %s ORDER BY trial_num",
            (robot_id,))

    def get_trial_values(self, robot_id, max_trials=5):
        """Returns list of raw values (strings/ints) for display, padded to max_trials."""
        trials = self.get_trials(robot_id, max_trials)
        result = [None] * max_trials
        for t in trials:
            idx = t["trial_num"] - 1
            if 0 <= idx < max_trials:
                result[idx] = t["value"]
        return result

    def next_trial_num(self, robot_id, max_trials=5):
        rows = self._q("SELECT MAX(trial_num) as mx FROM trials WHERE robot_id = %s", (robot_id,))
        mx = rows[0]["mx"] if rows and rows[0]["mx"] else 0
        nxt = mx + 1
        return nxt if nxt <= max_trials else None

    def write_trial_time(self, robot_id, time_str, max_trials=5):
        """Write a time trial (line following). Returns trial_num or None."""
        tn = self.next_trial_num(robot_id, max_trials)
        if tn is None:
            return None
        ms = parse_time(time_str)
        is_dnf = str(time_str).strip().upper() == "DNF"
        self._x(
            "INSERT INTO trials (robot_id, trial_num, value, value_ms, is_dnf) VALUES (%s,%s,%s,%s,%s)",
            (robot_id, tn, str(time_str), ms, is_dnf))
        return tn

    def write_trial_points(self, robot_id, points, max_trials=5):
        """Write a points trial (fire sister). Returns trial_num or None."""
        tn = self.next_trial_num(robot_id, max_trials)
        if tn is None:
            return None
        is_dnf = str(points).strip().upper() == "DNF"
        pts = None if is_dnf else int(points)
        self._x(
            "INSERT INTO trials (robot_id, trial_num, value, value_points, is_dnf) VALUES (%s,%s,%s,%s,%s)",
            (robot_id, tn, str(points), pts, is_dnf))
        return tn

    def best_time(self, robot_id):
        """Best time in ms (lowest) for line following."""
        rows = self._q(
            "SELECT MIN(value_ms) as best FROM trials WHERE robot_id = %s AND value_ms IS NOT NULL",
            (robot_id,))
        return rows[0]["best"] if rows and rows[0]["best"] is not None else None

    def best_points(self, robot_id):
        """Best points (highest) for fire sister."""
        rows = self._q(
            "SELECT MAX(value_points) as best FROM trials WHERE robot_id = %s AND value_points IS NOT NULL",
            (robot_id,))
        return rows[0]["best"] if rows and rows[0]["best"] is not None else None

    # ── Scoreboard ──

    def scoreboard_lf(self):
        """Full line following scoreboard."""
        robots = self.get_robots("line_following")
        board = []
        for r in robots:
            trials = self.get_trial_values(r["id"])
            best = self.best_time(r["id"])
            board.append({"num": r["number"], "name": r["name"], "id": r["id"],
                          "trials": trials, "best": best})
        board.sort(key=lambda x: (x["best"] is None, x["best"] or 0))
        return board

    def scoreboard_fs(self):
        """Full fire sister scoreboard."""
        robots = self.get_robots("fire_sister")
        board = []
        for r in robots:
            trials = self.get_trial_values(r["id"])
            best = self.best_points(r["id"])
            board.append({"num": r["number"], "name": r["name"], "id": r["id"],
                          "trials": trials, "best": best})
        board.sort(key=lambda x: (x["best"] is None, -(x["best"] or 0)))
        return board

    # ── Folkrace ──

    def add_folkrace_group(self, name, section="L", sort_order=0):
        self._x(
            "INSERT INTO folkrace_groups (name, section, sort_order) VALUES (%s, %s, %s)",
            (name, section, sort_order))
        rows = self._q("SELECT id FROM folkrace_groups WHERE name = %s ORDER BY id DESC LIMIT 1", (name,))
        return rows[0]["id"] if rows else None

    def get_folkrace_groups(self):
        return self._q("SELECT id, name, section FROM folkrace_groups ORDER BY sort_order, id")

    def add_folkrace_entry(self, group_id, robot_id, r1=0, r2=0, r3=0):
        total = r1 + r2 + r3
        self._x(
            "INSERT INTO folkrace_entries (group_id, robot_id, r1, r2, r3, total) VALUES (%s,%s,%s,%s,%s,%s)",
            (group_id, robot_id, r1, r2, r3, total))

    def get_folkrace_entries(self, group_id):
        return self._q("""
            SELECT fe.id, fe.robot_id, r.number, r.name, fe.r1, fe.r2, fe.r3, fe.total
            FROM folkrace_entries fe
            JOIN robots r ON r.id = fe.robot_id
            WHERE fe.group_id = %s
            ORDER BY fe.id
        """, (group_id,))

    def update_folkrace_entry(self, entry_id, r1, r2, r3):
        total = r1 + r2 + r3
        self._x("UPDATE folkrace_entries SET r1=%s, r2=%s, r3=%s, total=%s WHERE id=%s",
                (r1, r2, r3, total, entry_id))

    def get_folkrace_groups_full(self):
        """Return groups with robot entries for the GUI."""
        groups = self.get_folkrace_groups()
        for g in groups:
            g["robots"] = self.get_folkrace_entries(g["id"])
        return groups

    # ── Import from Excel ──

    def import_from_excel(self, filepath, cfg):
        """Import robots and existing data from the Excel tournament file."""
        if not HAS_XL:
            raise ImportError("pip install openpyxl")
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Line Following
        lf = cfg.get("lf", {})
        if lf.get("sheet") and lf["sheet"] in wb.sheetnames:
            ws = wb[lf["sheet"]]
            cn, cm, ct = lf.get("cn", "A"), lf.get("cm", "B"), lf.get("ct", "D")
            nt = lf.get("nt", 5)
            for r in range(lf.get("start", 7), ws.max_row + 1):
                num = ws[f"{cn}{r}"].value
                if num is None or not str(num).strip():
                    continue
                try: num_s = str(int(num))
                except: num_s = str(num).strip()
                name = str(ws[f"{cm}{r}"].value or "").strip()
                self.add_robot(num_s, name, "line_following")
                robot = self.find_robot(num_s, "line_following")
                if robot:
                    for i in range(nt):
                        v = ws[f"{chr(ord(ct)+i)}{r}"].value
                        if v is not None and str(v).strip():
                            self.write_trial_time(robot["id"], str(v), nt)

        # Fire Sister
        fs = cfg.get("fs", {})
        if fs.get("sheet") and fs["sheet"] in wb.sheetnames:
            ws = wb[fs["sheet"]]
            cn, cm, ct = fs.get("cn", "A"), fs.get("cm", "B"), fs.get("ct", "D")
            nt = fs.get("nt", 5)
            for r in range(fs.get("start", 4), ws.max_row + 1):
                num = ws[f"{cn}{r}"].value
                if num is None or not str(num).strip():
                    continue
                try: num_s = str(int(num))
                except: num_s = str(num).strip()
                name = str(ws[f"{cm}{r}"].value or "").strip()
                self.add_robot(num_s, name, "fire_sister")
                robot = self.find_robot(num_s, "fire_sister")
                if robot:
                    for i in range(nt):
                        v = ws[f"{chr(ord(ct)+i)}{r}"].value
                        if v is not None and str(v).strip():
                            if str(v).strip().upper() == "DNF":
                                self.write_trial_points(robot["id"], "DNF", nt)
                            else:
                                try:
                                    self.write_trial_points(robot["id"], int(float(str(v))), nt)
                                except:
                                    pass

        # Folkrace
        fr = cfg.get("fr", {})
        if fr.get("sheet") and fr["sheet"] in wb.sheetnames:
            ws = wb[fr["sheet"]]
            # Import robot list
            ln, lm, ls = fr.get("ln", "P"), fr.get("lm", "Q"), fr.get("ls", 6)
            for r in range(ls, ws.max_row + 1):
                num = ws[f"{ln}{r}"].value
                if num is None or not str(num).strip():
                    continue
                try: num_s = str(int(num))
                except: num_s = str(num).strip()
                name = str(ws[f"{lm}{r}"].value or "").strip()
                self.add_robot(num_s, name, "folkrace")

            # Import groups (left section)
            gc_col = fr.get("gc", "B")
            gn_col = fr.get("gn", "C")
            gr_cols = fr.get("gr", ["D", "E", "F"])
            gt_col = fr.get("gt", "G")
            cur_gid = None
            sort_idx = 0
            for r in range(1, ws.max_row + 1):
                v = ws[f"{gc_col}{r}"].value
                if v and isinstance(v, str) and any(k in v for k in ["Grup", "Group"]):
                    cur_gid = self.add_folkrace_group(v.strip(), "L", sort_idx)
                    sort_idx += 1
                elif cur_gid and v is not None and str(v).strip():
                    try: num = str(int(float(str(v))))
                    except: continue
                    robot = self.find_robot(num, "folkrace")
                    if robot:
                        rs = []
                        for c in gr_cols:
                            sv = ws[f"{c}{r}"].value
                            try: rs.append(int(sv) if sv is not None else 0)
                            except: rs.append(0)
                        while len(rs) < 3: rs.append(0)
                        self.add_folkrace_entry(cur_gid, robot["id"], rs[0], rs[1], rs[2])

            # Right section (Semi/Finals)
            cur_gid = None
            for r in range(1, ws.max_row + 1):
                v = ws[f"I{r}"].value
                if v and isinstance(v, str) and any(k in str(v) for k in ["Semi", "Final", "Pusfinalis"]):
                    cur_gid = self.add_folkrace_group(str(v).strip(), "R", sort_idx)
                    sort_idx += 1
                elif cur_gid and v is not None:
                    try: num = str(int(float(str(v))))
                    except: continue
                    robot = self.find_robot(num, "folkrace")
                    if robot:
                        rs = []
                        for c in ["K", "L", "M"]:
                            sv = ws[f"{c}{r}"].value
                            try: rs.append(int(sv) if sv is not None else 0)
                            except: rs.append(0)
                        while len(rs) < 3: rs.append(0)
                        self.add_folkrace_entry(cur_gid, robot["id"], rs[0], rs[1], rs[2])

        count_r = self._q("SELECT COUNT(*) as c FROM robots")
        count_t = self._q("SELECT COUNT(*) as c FROM trials")
        count_g = self._q("SELECT COUNT(*) as c FROM folkrace_groups")
        return {
            "robots": count_r[0]["c"],
            "trials": count_t[0]["c"],
            "groups": count_g[0]["c"]
        }


class GSheetSync:
    """Optional background sync: pushes DB data to a Google Sheet periodically."""

    def __init__(self, db, creds_file, sheet_id, interval=30):
        self.db = db
        self.interval = interval
        self._stop = False
        self.gs = None
        self.error = None

        if not HAS_GS:
            self.error = "gspread not installed"
            return

        try:
            creds = Credentials.from_service_account_file(creds_file,
                scopes=["https://www.googleapis.com/auth/spreadsheets",
                         "https://www.googleapis.com/auth/drive"])
            gc = gspread.authorize(creds)
            s = sheet_id.strip()
            if "/d/" in s:
                s = s.split("/d/")[1].split("/")[0]
            self.gs = gc.open_by_key(s)
            print(f"[GSheets Sync] Connected to: {self.gs.title}")
        except Exception as e:
            self.error = str(e)
            print(f"[GSheets Sync] Failed: {e}")
            return

        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop = True

    def _loop(self):
        while not self._stop:
            try:
                self._sync()
            except Exception as e:
                print(f"[GSheets Sync] Error: {e}")
            time.sleep(self.interval)

    def _sync(self):
        if not self.gs or not self.db.ok:
            return

        # Sync line following
        try:
            board = self.db.scoreboard_lf()
            ws = self._get_ws("Line Following")
            data = [["#", "Name", "T1", "T2", "T3", "T4", "T5", "Best"]]
            for r in board:
                row = [r["num"], r["name"]]
                for t in r["trials"]:
                    row.append(str(t) if t is not None else "")
                best = r["best"]
                row.append(fmt_ms_sync(best) if best else "")
                data.append(row)
            ws.clear()
            if data:
                ws.update(data, "A1")
        except Exception as e:
            print(f"[GSheets Sync] LF error: {e}")

        # Sync fire sister
        try:
            board = self.db.scoreboard_fs()
            ws = self._get_ws("Fire Sister")
            data = [["#", "Name", "T1", "T2", "T3", "T4", "T5", "Best"]]
            for r in board:
                row = [r["num"], r["name"]]
                for t in r["trials"]:
                    row.append(str(t) if t is not None else "")
                best = r["best"]
                row.append(f"{best} pts" if best is not None else "")
                data.append(row)
            ws.clear()
            if data:
                ws.update(data, "A1")
        except Exception as e:
            print(f"[GSheets Sync] FS error: {e}")

        # Sync folkrace
        try:
            groups = self.db.get_folkrace_groups_full()
            ws = self._get_ws("Folkrace")
            data = []
            for g in groups:
                data.append([g["name"]])
                data.append(["#", "Robot", "R1", "R2", "R3", "Total"])
                for ent in g["robots"]:
                    data.append([ent["number"], ent["name"],
                                ent["r1"], ent["r2"], ent["r3"], ent["total"]])
                data.append([])
            ws.clear()
            if data:
                ws.update(data, "A1")
        except Exception as e:
            print(f"[GSheets Sync] FR error: {e}")

    def _get_ws(self, name):
        try:
            return self.gs.worksheet(name)
        except:
            return self.gs.add_worksheet(title=name, rows=200, cols=10)


def fmt_ms_sync(ms):
    if ms is None:
        return ""
    ms = max(0, int(ms))
    m, s, ml = ms // 60000, (ms % 60000) // 1000, ms % 1000
    return f"{m}:{s:02d}.{ml:03d}" if m else f"{s}.{ml:03d}"
